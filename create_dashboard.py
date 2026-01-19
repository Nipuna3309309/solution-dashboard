import pandas as pd
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment, GradientFill
from openpyxl.chart import BarChart, PieChart, LineChart, Reference, DoughnutChart
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.utils import get_column_letter
from copy import copy
import warnings
warnings.filterwarnings('ignore')

# Color palette (Power BI inspired)
COLORS = {
    'primary': '118DFF',      # Blue
    'secondary': '12239E',    # Dark Blue
    'accent1': 'E66C37',      # Orange
    'accent2': '6B007B',      # Purple
    'accent3': 'E044A7',      # Pink
    'accent4': '744EC2',      # Violet
    'accent5': 'D9B300',      # Yellow
    'success': '30B177',      # Green
    'background': 'F2F2F2',   # Light Gray
    'dark_bg': '252423',      # Dark background
    'card_bg': 'FFFFFF',      # White
    'text_dark': '252423',
    'text_light': '666666',
}

def create_kpi_card(ws, start_row, start_col, title, formula, color, width=3):
    """Create a KPI card with title and value"""
    # Merge cells for card
    end_col = start_col + width - 1

    # Card background
    card_fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )

    # Top colored bar
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=start_row, column=col)
        cell.fill = card_fill
        cell.border = thin_border
    ws.row_dimensions[start_row].height = 8

    # Title row
    ws.merge_cells(start_row=start_row+1, start_column=start_col, end_row=start_row+1, end_column=end_col)
    title_cell = ws.cell(row=start_row+1, column=start_col)
    title_cell.value = title
    title_cell.font = Font(name='Segoe UI', size=10, color='666666')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.fill = white_fill
    ws.row_dimensions[start_row+1].height = 25

    for col in range(start_col, end_col + 1):
        ws.cell(row=start_row+1, column=col).border = thin_border
        ws.cell(row=start_row+1, column=col).fill = white_fill

    # Value row
    ws.merge_cells(start_row=start_row+2, start_column=start_col, end_row=start_row+2, end_column=end_col)
    value_cell = ws.cell(row=start_row+2, column=start_col)
    value_cell.value = formula
    value_cell.font = Font(name='Segoe UI Semibold', size=24, color=color, bold=True)
    value_cell.alignment = Alignment(horizontal='center', vertical='center')
    value_cell.fill = white_fill
    ws.row_dimensions[start_row+2].height = 50

    for col in range(start_col, end_col + 1):
        ws.cell(row=start_row+2, column=col).border = thin_border
        ws.cell(row=start_row+2, column=col).fill = white_fill

    # Bottom padding
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=start_row+3, column=col)
        cell.fill = white_fill
        cell.border = thin_border
    ws.row_dimensions[start_row+3].height = 10

def create_section_header(ws, row, col, end_col, title):
    """Create a section header"""
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=end_col)
    cell = ws.cell(row=row, column=col)
    cell.value = title
    cell.font = Font(name='Segoe UI Semibold', size=14, color='252423', bold=True)
    cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[row].height = 35

def main():
    # Load existing data
    df = pd.read_excel('Solution List.xlsx')
    df.columns = df.columns.str.strip()  # Remove whitespace from column names

    # Fill NaN with 0 for calculations
    df_calc = df.copy()
    df_calc['SMV Unlock'] = df_calc['SMV Unlock'].fillna(0)
    df_calc['OH Reduction'] = df_calc['OH Reduction'].fillna(0)
    df_calc['Other Savings'] = df_calc['Other Savings'].fillna(0)

    # Create workbook
    wb = Workbook()

    # ========== DATA SHEET ==========
    ws_data = wb.active
    ws_data.title = "Data"

    # Write data with formatting
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_data.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='118DFF', end_color='118DFF', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.font = Font(name='Segoe UI', size=10)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                bottom=Side(style='thin', color='DDDDDD')
            )

    # Auto-fit columns
    for col in range(1, 8):
        ws_data.column_dimensions[get_column_letter(col)].width = 18

    # ========== PIVOT DATA SHEET (for charts) ==========
    ws_pivot = wb.create_sheet("PivotData")

    # Division Summary
    div_summary = df_calc.groupby('Division').agg({
        'SMV Unlock': 'sum',
        'OH Reduction': 'sum',
        'Other Savings': 'sum',
        'Solution Name': 'count'
    }).reset_index()
    div_summary.columns = ['Division', 'SMV Unlock', 'OH Reduction', 'Other Savings', 'Solution Count']
    div_summary['Total Savings'] = div_summary['SMV Unlock'] + div_summary['OH Reduction'] + div_summary['Other Savings']

    ws_pivot['A1'] = 'DIVISION SUMMARY'
    ws_pivot['A1'].font = Font(bold=True, size=12)
    for r_idx, row in enumerate(dataframe_to_rows(div_summary, index=False, header=True), 3):
        for c_idx, value in enumerate(row, 1):
            ws_pivot.cell(row=r_idx, column=c_idx, value=value)

    div_end_row = 3 + len(div_summary)

    # Stage Summary
    stage_summary = df_calc.groupby('Stage').agg({
        'SMV Unlock': 'sum',
        'OH Reduction': 'sum',
        'Other Savings': 'sum',
        'Solution Name': 'count'
    }).reset_index()
    stage_summary.columns = ['Stage', 'SMV Unlock', 'OH Reduction', 'Other Savings', 'Solution Count']

    stage_start = div_end_row + 3
    ws_pivot.cell(row=stage_start, column=1, value='STAGE SUMMARY').font = Font(bold=True, size=12)
    for r_idx, row in enumerate(dataframe_to_rows(stage_summary, index=False, header=True), stage_start + 2):
        for c_idx, value in enumerate(row, 1):
            ws_pivot.cell(row=r_idx, column=c_idx, value=value)

    stage_end_row = stage_start + 2 + len(stage_summary)

    # Focus Area Summary
    focus_summary = df_calc.groupby('Focus Area').agg({
        'SMV Unlock': 'sum',
        'OH Reduction': 'sum',
        'Other Savings': 'sum',
        'Solution Name': 'count'
    }).reset_index()
    focus_summary.columns = ['Focus Area', 'SMV Unlock', 'OH Reduction', 'Other Savings', 'Solution Count']

    focus_start = stage_end_row + 3
    ws_pivot.cell(row=focus_start, column=1, value='FOCUS AREA SUMMARY').font = Font(bold=True, size=12)
    for r_idx, row in enumerate(dataframe_to_rows(focus_summary, index=False, header=True), focus_start + 2):
        for c_idx, value in enumerate(row, 1):
            ws_pivot.cell(row=focus_start + 2 + r_idx - (focus_start + 2), column=c_idx, value=value)

    # Actually write focus area data properly
    for r_idx, row in enumerate(dataframe_to_rows(focus_summary, index=False, header=True), focus_start + 2):
        for c_idx, value in enumerate(row, 1):
            ws_pivot.cell(row=r_idx, column=c_idx, value=value)

    focus_end_row = focus_start + 2 + len(focus_summary)

    # ========== DASHBOARD SHEET ==========
    ws_dash = wb.create_sheet("Dashboard")
    wb.active = ws_dash

    # Set column widths
    for col in range(1, 20):
        ws_dash.column_dimensions[get_column_letter(col)].width = 12

    # Background color
    gray_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
    for row in range(1, 60):
        for col in range(1, 20):
            ws_dash.cell(row=row, column=col).fill = gray_fill

    # ========== HEADER ==========
    ws_dash.merge_cells('B2:R2')
    header_cell = ws_dash['B2']
    header_cell.value = "SOLUTION SAVINGS DASHBOARD"
    header_cell.font = Font(name='Segoe UI', size=28, bold=True, color='118DFF')
    header_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws_dash.row_dimensions[2].height = 50

    # Subtitle
    ws_dash.merge_cells('B3:R3')
    sub_cell = ws_dash['B3']
    sub_cell.value = "Real-time Analytics | Division Performance | Savings Breakdown"
    sub_cell.font = Font(name='Segoe UI', size=11, color='666666', italic=True)
    sub_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws_dash.row_dimensions[3].height = 25

    # ========== KPI CARDS (Row 5-8) ==========
    # Total Solutions
    create_kpi_card(ws_dash, 5, 2, "TOTAL SOLUTIONS",
                   f"=COUNTA(Data!B2:B{len(df)+1})", COLORS['primary'], 3)

    # Total SMV Unlock
    create_kpi_card(ws_dash, 5, 6, "TOTAL SMV UNLOCK",
                   f"=ROUND(SUM(Data!E2:E{len(df)+1}),3)", COLORS['success'], 3)

    # Total OH Reduction
    create_kpi_card(ws_dash, 5, 10, "TOTAL OH REDUCTION",
                   f"=ROUND(SUM(Data!F2:F{len(df)+1}),1)", COLORS['accent1'], 3)

    # Total Other Savings
    create_kpi_card(ws_dash, 5, 14, "OTHER SAVINGS",
                   f"=ROUND(SUM(Data!G2:G{len(df)+1}),1)", COLORS['accent2'], 3)

    # ========== CHARTS SECTION ==========

    # --- Division Performance Bar Chart ---
    create_section_header(ws_dash, 11, 2, 8, "Division Performance")

    chart1 = BarChart()
    chart1.type = "col"
    chart1.grouping = "clustered"
    chart1.title = None
    chart1.style = 10

    # Data references for division chart
    data_ref = Reference(ws_pivot, min_col=2, min_row=3, max_col=4, max_row=3+len(div_summary))
    cats_ref = Reference(ws_pivot, min_col=1, min_row=4, max_row=3+len(div_summary))

    chart1.add_data(data_ref, titles_from_data=True)
    chart1.set_categories(cats_ref)
    chart1.shape = 4
    chart1.width = 18
    chart1.height = 10

    # Color the series
    colors_list = ['118DFF', '30B177', 'E66C37']
    for i, series in enumerate(chart1.series):
        series.graphicalProperties.solidFill = colors_list[i % len(colors_list)]

    chart1.legend.position = 'b'
    ws_dash.add_chart(chart1, "B12")

    # --- Stage Distribution Doughnut Chart ---
    create_section_header(ws_dash, 11, 11, 17, "Stage Distribution")

    chart2 = DoughnutChart()
    chart2.title = None
    chart2.style = 10

    stage_data_ref = Reference(ws_pivot, min_col=5, min_row=stage_start+2, max_row=stage_start+2+len(stage_summary))
    stage_cats_ref = Reference(ws_pivot, min_col=1, min_row=stage_start+3, max_row=stage_start+2+len(stage_summary))

    chart2.add_data(stage_data_ref, titles_from_data=True)
    chart2.set_categories(stage_cats_ref)
    chart2.width = 14
    chart2.height = 10
    chart2.holeSize = 50

    # Add data labels
    chart2.dataLabels = DataLabelList()
    chart2.dataLabels.showPercent = True
    chart2.dataLabels.showCatName = True
    chart2.dataLabels.showVal = False

    ws_dash.add_chart(chart2, "K12")

    # --- Focus Area Analysis ---
    create_section_header(ws_dash, 28, 2, 8, "Focus Area Analysis")

    chart3 = BarChart()
    chart3.type = "bar"
    chart3.grouping = "stacked"
    chart3.title = None
    chart3.style = 10

    focus_data_ref = Reference(ws_pivot, min_col=2, min_row=focus_start+2, max_col=4, max_row=focus_start+2+len(focus_summary))
    focus_cats_ref = Reference(ws_pivot, min_col=1, min_row=focus_start+3, max_row=focus_start+2+len(focus_summary))

    chart3.add_data(focus_data_ref, titles_from_data=True)
    chart3.set_categories(focus_cats_ref)
    chart3.width = 18
    chart3.height = 10

    for i, series in enumerate(chart3.series):
        series.graphicalProperties.solidFill = colors_list[i % len(colors_list)]

    chart3.legend.position = 'b'
    ws_dash.add_chart(chart3, "B29")

    # --- Total Savings by Division (Pie) ---
    create_section_header(ws_dash, 28, 11, 17, "Total Savings by Division")

    chart4 = PieChart()
    chart4.title = None
    chart4.style = 10

    total_sav_ref = Reference(ws_pivot, min_col=6, min_row=3, max_row=3+len(div_summary))
    div_cats_ref = Reference(ws_pivot, min_col=1, min_row=4, max_row=3+len(div_summary))

    chart4.add_data(total_sav_ref, titles_from_data=True)
    chart4.set_categories(div_cats_ref)
    chart4.width = 14
    chart4.height = 10

    chart4.dataLabels = DataLabelList()
    chart4.dataLabels.showPercent = True
    chart4.dataLabels.showCatName = True

    ws_dash.add_chart(chart4, "K29")

    # ========== TOP PERFORMERS TABLE ==========
    create_section_header(ws_dash, 45, 2, 10, "Top Performing Divisions")

    # Sort divisions by total savings
    div_sorted = div_summary.sort_values('Total Savings', ascending=False)

    # Table headers
    headers = ['Rank', 'Division', 'SMV Unlock', 'OH Reduction', 'Other Savings', 'Total']
    header_fill = PatternFill(start_color='118DFF', end_color='118DFF', fill_type='solid')

    for i, header in enumerate(headers):
        cell = ws_dash.cell(row=46, column=2+i, value=header)
        cell.font = Font(name='Segoe UI', size=10, bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )

    ws_dash.row_dimensions[46].height = 25

    # Table data with conditional formatting colors
    rank_colors = ['FFD700', 'C0C0C0', 'CD7F32', 'FFFFFF', 'FFFFFF']  # Gold, Silver, Bronze

    for idx, (_, row) in enumerate(div_sorted.iterrows()):
        row_num = 47 + idx
        data = [idx + 1, row['Division'], round(row['SMV Unlock'], 3),
                round(row['OH Reduction'], 1), round(row['Other Savings'], 1),
                round(row['Total Savings'], 2)]

        row_fill = PatternFill(start_color=rank_colors[min(idx, 4)],
                               end_color=rank_colors[min(idx, 4)], fill_type='solid')

        for i, value in enumerate(data):
            cell = ws_dash.cell(row=row_num, column=2+i, value=value)
            cell.font = Font(name='Segoe UI', size=10,
                           bold=(idx < 3),
                           color='252423' if idx < 3 else '666666')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if idx < 3:
                cell.fill = row_fill
            cell.border = Border(
                left=Side(style='thin', color='DDDDDD'),
                right=Side(style='thin', color='DDDDDD'),
                top=Side(style='thin', color='DDDDDD'),
                bottom=Side(style='thin', color='DDDDDD')
            )
        ws_dash.row_dimensions[row_num].height = 22

    # ========== INSIGHTS PANEL ==========
    create_section_header(ws_dash, 45, 11, 17, "Key Insights")

    # Calculate insights
    top_div = div_sorted.iloc[0]['Division']
    top_smv_div = div_summary.loc[div_summary['SMV Unlock'].idxmax(), 'Division']
    top_oh_div = div_summary.loc[div_summary['OH Reduction'].idxmax(), 'Division']
    total_solutions = len(df)

    insights = [
        f"Top Performer: {top_div}",
        f"Best SMV Unlock: {top_smv_div}",
        f"Best OH Reduction: {top_oh_div}",
        f"Total Solutions: {total_solutions}",
        f"Divisions Tracked: {len(div_summary)}"
    ]

    insight_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

    for idx, insight in enumerate(insights):
        row_num = 46 + idx
        ws_dash.merge_cells(start_row=row_num, start_column=11, end_row=row_num, end_column=17)
        cell = ws_dash.cell(row=row_num, column=11, value=f"  {insight}")
        cell.font = Font(name='Segoe UI', size=11, color='252423')
        cell.fill = insight_fill
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = Border(
            left=Side(style='medium', color='118DFF'),
            bottom=Side(style='thin', color='EEEEEE')
        )
        ws_dash.row_dimensions[row_num].height = 28

    # ========== FOOTER ==========
    ws_dash.merge_cells('B55:R55')
    footer = ws_dash['B55']
    footer.value = "Data automatically updates from 'Data' sheet | Dashboard created with Python & OpenPyXL"
    footer.font = Font(name='Segoe UI', size=9, color='999999', italic=True)
    footer.alignment = Alignment(horizontal='center', vertical='center')

    # Hide gridlines on dashboard
    ws_dash.sheet_view.showGridLines = False

    # Set print area
    ws_dash.print_area = 'A1:S56'

    # Move Dashboard to first position
    wb.move_sheet(ws_dash, offset=-2)

    # Save workbook
    wb.save('Solution_Dashboard.xlsx')
    print("Dashboard created successfully: Solution_Dashboard.xlsx")
    print(f"\nSummary Statistics:")
    print(f"  Total Solutions: {len(df)}")
    print(f"  Total SMV Unlock: {df_calc['SMV Unlock'].sum():.3f}")
    print(f"  Total OH Reduction: {df_calc['OH Reduction'].sum():.1f}")
    print(f"  Total Other Savings: {df_calc['Other Savings'].sum():.1f}")
    print(f"\nTop Division: {top_div}")
    print(f"Best SMV Unlock Division: {top_smv_div}")
    print(f"Best OH Reduction Division: {top_oh_div}")

if __name__ == "__main__":
    main()
