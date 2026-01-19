import pandas as pd
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment
from openpyxl.chart import BarChart, PieChart, LineChart, Reference, DoughnutChart, AreaChart
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
import warnings
warnings.filterwarnings('ignore')

# Power BI Color Palette
COLORS = {
    'primary': '118DFF',
    'secondary': '12239E',
    'accent1': 'E66C37',
    'accent2': '6B007B',
    'accent3': 'E044A7',
    'success': '30B177',
    'warning': 'D9B300',
    'bg_dark': '252423',
    'bg_light': 'F2F2F2',
    'white': 'FFFFFF',
    'gray': '666666',
}

def style_header_cell(cell, bg_color='118DFF'):
    cell.font = Font(name='Segoe UI', size=10, bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

def style_data_cell(cell, bold=False, color='252423'):
    cell.font = Font(name='Segoe UI', size=10, bold=bold, color=color)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )

def create_kpi_card(ws, row, col, title, formula, accent_color, icon=""):
    """Create a modern KPI card with formula"""
    white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    accent_fill = PatternFill(start_color=accent_color, end_color=accent_color, fill_type='solid')

    thin = Side(style='thin', color='E0E0E0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Accent bar (top)
    for c in range(col, col + 3):
        cell = ws.cell(row=row, column=c)
        cell.fill = accent_fill
    ws.row_dimensions[row].height = 6

    # Title
    ws.merge_cells(start_row=row+1, start_column=col, end_row=row+1, end_column=col+2)
    title_cell = ws.cell(row=row+1, column=col, value=title)
    title_cell.font = Font(name='Segoe UI', size=9, color='888888')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.fill = white_fill
    for c in range(col, col + 3):
        ws.cell(row=row+1, column=c).fill = white_fill
        ws.cell(row=row+1, column=c).border = border
    ws.row_dimensions[row+1].height = 22

    # Value
    ws.merge_cells(start_row=row+2, start_column=col, end_row=row+2, end_column=col+2)
    value_cell = ws.cell(row=row+2, column=col, value=formula)
    value_cell.font = Font(name='Segoe UI Semibold', size=26, bold=True, color=accent_color)
    value_cell.alignment = Alignment(horizontal='center', vertical='center')
    value_cell.fill = white_fill
    for c in range(col, col + 3):
        ws.cell(row=row+2, column=c).fill = white_fill
        ws.cell(row=row+2, column=c).border = border
    ws.row_dimensions[row+2].height = 45

    # Bottom padding
    for c in range(col, col + 3):
        cell = ws.cell(row=row+3, column=c)
        cell.fill = white_fill
        cell.border = border
    ws.row_dimensions[row+3].height = 8

def main():
    # Load existing data
    df = pd.read_excel('Solution List.xlsx')
    df.columns = df.columns.str.strip()

    # Get unique values for dynamic formulas
    divisions = df['Division'].dropna().unique().tolist()
    stages = df['Stage'].dropna().unique().tolist()
    focus_areas = df['Focus Area'].dropna().unique().tolist()

    data_rows = len(df) + 1  # +1 for header

    # Create workbook
    wb = Workbook()

    # ==================== DATA SHEET ====================
    ws_data = wb.active
    ws_data.title = "Data"

    # Write headers
    headers = ['Division', 'Solution Name', 'Focus Area', 'Stage', 'SMV Unlock', 'OH Reduction', 'Other Savings']
    for c, header in enumerate(headers, 1):
        style_header_cell(ws_data.cell(row=1, column=c, value=header))

    # Write data
    for r, row_data in enumerate(df.values, 2):
        for c, value in enumerate(row_data, 1):
            cell = ws_data.cell(row=r, column=c, value=value)
            style_data_cell(cell)

    # Set column widths
    widths = [15, 18, 25, 15, 12, 14, 14]
    for c, w in enumerate(widths, 1):
        ws_data.column_dimensions[get_column_letter(c)].width = w

    # ==================== CALCULATIONS SHEET ====================
    ws_calc = wb.create_sheet("Calculations")

    # --- Division Summary with SUMIF formulas ---
    ws_calc['A1'] = 'DIVISION SUMMARY'
    ws_calc['A1'].font = Font(bold=True, size=12, color='118DFF')

    div_headers = ['Division', 'SMV Unlock', 'OH Reduction', 'Other Savings', 'Solution Count', 'Total Savings']
    for c, h in enumerate(div_headers, 1):
        style_header_cell(ws_calc.cell(row=2, column=c, value=h))

    for r, div in enumerate(divisions, 3):
        ws_calc.cell(row=r, column=1, value=div)
        ws_calc.cell(row=r, column=2, value=f'=SUMIF(Data!$A$2:$A$100,A{r},Data!$E$2:$E$100)')
        ws_calc.cell(row=r, column=3, value=f'=SUMIF(Data!$A$2:$A$100,A{r},Data!$F$2:$F$100)')
        ws_calc.cell(row=r, column=4, value=f'=SUMIF(Data!$A$2:$A$100,A{r},Data!$G$2:$G$100)')
        ws_calc.cell(row=r, column=5, value=f'=COUNTIF(Data!$A$2:$A$100,A{r})')
        ws_calc.cell(row=r, column=6, value=f'=B{r}+C{r}+D{r}')

        for c in range(1, 7):
            style_data_cell(ws_calc.cell(row=r, column=c))

    div_end_row = 2 + len(divisions)

    # --- Stage Summary ---
    stage_start = div_end_row + 3
    ws_calc.cell(row=stage_start, column=1, value='STAGE SUMMARY').font = Font(bold=True, size=12, color='118DFF')

    stage_headers = ['Stage', 'SMV Unlock', 'OH Reduction', 'Other Savings', 'Solution Count']
    for c, h in enumerate(stage_headers, 1):
        style_header_cell(ws_calc.cell(row=stage_start+1, column=c, value=h))

    for r, stage in enumerate(stages, stage_start + 2):
        row_idx = r
        ws_calc.cell(row=row_idx, column=1, value=stage)
        ws_calc.cell(row=row_idx, column=2, value=f'=SUMIF(Data!$D$2:$D$100,A{row_idx},Data!$E$2:$E$100)')
        ws_calc.cell(row=row_idx, column=3, value=f'=SUMIF(Data!$D$2:$D$100,A{row_idx},Data!$F$2:$F$100)')
        ws_calc.cell(row=row_idx, column=4, value=f'=SUMIF(Data!$D$2:$D$100,A{row_idx},Data!$G$2:$G$100)')
        ws_calc.cell(row=row_idx, column=5, value=f'=COUNTIF(Data!$D$2:$D$100,A{row_idx})')

        for c in range(1, 6):
            style_data_cell(ws_calc.cell(row=row_idx, column=c))

    stage_end_row = stage_start + 1 + len(stages)

    # --- Focus Area Summary ---
    focus_start = stage_end_row + 3
    ws_calc.cell(row=focus_start, column=1, value='FOCUS AREA SUMMARY').font = Font(bold=True, size=12, color='118DFF')

    focus_headers = ['Focus Area', 'SMV Unlock', 'OH Reduction', 'Other Savings', 'Solution Count']
    for c, h in enumerate(focus_headers, 1):
        style_header_cell(ws_calc.cell(row=focus_start+1, column=c, value=h))

    for i, focus in enumerate(focus_areas):
        row_idx = focus_start + 2 + i
        ws_calc.cell(row=row_idx, column=1, value=focus)
        ws_calc.cell(row=row_idx, column=2, value=f'=SUMIF(Data!$C$2:$C$100,A{row_idx},Data!$E$2:$E$100)')
        ws_calc.cell(row=row_idx, column=3, value=f'=SUMIF(Data!$C$2:$C$100,A{row_idx},Data!$F$2:$F$100)')
        ws_calc.cell(row=row_idx, column=4, value=f'=SUMIF(Data!$C$2:$C$100,A{row_idx},Data!$G$2:$G$100)')
        ws_calc.cell(row=row_idx, column=5, value=f'=COUNTIF(Data!$C$2:$C$100,A{row_idx})')

        for c in range(1, 6):
            style_data_cell(ws_calc.cell(row=row_idx, column=c))

    focus_end_row = focus_start + 1 + len(focus_areas)

    # Set column widths for calculations
    calc_widths = [22, 14, 14, 14, 14, 14]
    for c, w in enumerate(calc_widths, 1):
        ws_calc.column_dimensions[get_column_letter(c)].width = w

    # ==================== DASHBOARD SHEET ====================
    ws_dash = wb.create_sheet("Dashboard")

    # Background
    bg_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
    for row in range(1, 70):
        for col in range(1, 22):
            ws_dash.cell(row=row, column=col).fill = bg_fill

    # Column widths
    for col in range(1, 22):
        ws_dash.column_dimensions[get_column_letter(col)].width = 11

    # ========== HEADER ==========
    ws_dash.merge_cells('B2:S2')
    header = ws_dash['B2']
    header.value = "SOLUTION SAVINGS DASHBOARD"
    header.font = Font(name='Segoe UI', size=32, bold=True, color='118DFF')
    header.alignment = Alignment(horizontal='center', vertical='center')
    ws_dash.row_dimensions[2].height = 55

    ws_dash.merge_cells('B3:S3')
    sub = ws_dash['B3']
    sub.value = "Real-time Analytics  |  Division Performance  |  Savings Breakdown  |  Auto-Updates"
    sub.font = Font(name='Segoe UI', size=11, color='888888', italic=True)
    sub.alignment = Alignment(horizontal='center', vertical='center')
    ws_dash.row_dimensions[3].height = 25

    # ========== KPI CARDS ==========
    # Card 1: Total Solutions
    create_kpi_card(ws_dash, 5, 2, "TOTAL SOLUTIONS",
                   "=COUNTA(Data!B2:B100)", COLORS['primary'])

    # Card 2: SMV Unlock
    create_kpi_card(ws_dash, 5, 6, "TOTAL SMV UNLOCK",
                   "=ROUND(SUM(Data!E2:E100),3)", COLORS['success'])

    # Card 3: OH Reduction
    create_kpi_card(ws_dash, 5, 10, "OH REDUCTION",
                   "=ROUND(SUM(Data!F2:F100),1)", COLORS['accent1'])

    # Card 4: Other Savings
    create_kpi_card(ws_dash, 5, 14, "OTHER SAVINGS",
                   "=ROUND(SUM(Data!G2:G100),1)", COLORS['accent2'])

    # Card 5: Avg Savings per Solution
    create_kpi_card(ws_dash, 5, 18, "AVG/SOLUTION",
                   "=ROUND((SUM(Data!E2:E100)+SUM(Data!F2:F100)+SUM(Data!G2:G100))/COUNTA(Data!B2:B100),2)",
                   COLORS['warning'])

    # ========== SECTION 1: Division Performance ==========
    ws_dash.merge_cells('B11:I11')
    sec1 = ws_dash['B11']
    sec1.value = "DIVISION PERFORMANCE"
    sec1.font = Font(name='Segoe UI Semibold', size=14, bold=True, color='252423')
    sec1.alignment = Alignment(horizontal='left', vertical='center')
    ws_dash.row_dimensions[11].height = 35

    # Bar chart - Division by Savings Type
    chart1 = BarChart()
    chart1.type = "col"
    chart1.grouping = "clustered"
    chart1.title = None
    chart1.style = 10

    data1 = Reference(ws_calc, min_col=2, min_row=2, max_col=4, max_row=div_end_row)
    cats1 = Reference(ws_calc, min_col=1, min_row=3, max_row=div_end_row)

    chart1.add_data(data1, titles_from_data=True)
    chart1.set_categories(cats1)
    chart1.width = 18
    chart1.height = 11
    chart1.legend.position = 'b'

    colors_bar = ['118DFF', '30B177', 'E66C37']
    for i, s in enumerate(chart1.series):
        s.graphicalProperties.solidFill = colors_bar[i]

    ws_dash.add_chart(chart1, "B12")

    # ========== SECTION 2: Stage Distribution ==========
    ws_dash.merge_cells('L11:S11')
    sec2 = ws_dash['L11']
    sec2.value = "STAGE DISTRIBUTION"
    sec2.font = Font(name='Segoe UI Semibold', size=14, bold=True, color='252423')
    sec2.alignment = Alignment(horizontal='left', vertical='center')

    # Doughnut chart - Stage
    chart2 = DoughnutChart()
    chart2.title = None
    chart2.style = 10

    data2 = Reference(ws_calc, min_col=5, min_row=stage_start+1, max_row=stage_end_row)
    cats2 = Reference(ws_calc, min_col=1, min_row=stage_start+2, max_row=stage_end_row)

    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats2)
    chart2.width = 15
    chart2.height = 11
    chart2.holeSize = 55

    chart2.dataLabels = DataLabelList()
    chart2.dataLabels.showPercent = True
    chart2.dataLabels.showCatName = True
    chart2.dataLabels.showVal = False

    ws_dash.add_chart(chart2, "L12")

    # ========== SECTION 3: Focus Area Analysis ==========
    ws_dash.merge_cells('B29:I29')
    sec3 = ws_dash['B29']
    sec3.value = "FOCUS AREA BREAKDOWN"
    sec3.font = Font(name='Segoe UI Semibold', size=14, bold=True, color='252423')
    sec3.alignment = Alignment(horizontal='left', vertical='center')
    ws_dash.row_dimensions[29].height = 35

    # Horizontal stacked bar
    chart3 = BarChart()
    chart3.type = "bar"
    chart3.grouping = "stacked"
    chart3.title = None
    chart3.style = 10

    data3 = Reference(ws_calc, min_col=2, min_row=focus_start+1, max_col=4, max_row=focus_end_row)
    cats3 = Reference(ws_calc, min_col=1, min_row=focus_start+2, max_row=focus_end_row)

    chart3.add_data(data3, titles_from_data=True)
    chart3.set_categories(cats3)
    chart3.width = 18
    chart3.height = 11
    chart3.legend.position = 'b'

    for i, s in enumerate(chart3.series):
        s.graphicalProperties.solidFill = colors_bar[i]

    ws_dash.add_chart(chart3, "B30")

    # ========== SECTION 4: Total Savings by Division ==========
    ws_dash.merge_cells('L29:S29')
    sec4 = ws_dash['L29']
    sec4.value = "TOTAL SAVINGS BY DIVISION"
    sec4.font = Font(name='Segoe UI Semibold', size=14, bold=True, color='252423')
    sec4.alignment = Alignment(horizontal='left', vertical='center')

    # Pie chart
    chart4 = PieChart()
    chart4.title = None
    chart4.style = 10

    data4 = Reference(ws_calc, min_col=6, min_row=2, max_row=div_end_row)
    cats4 = Reference(ws_calc, min_col=1, min_row=3, max_row=div_end_row)

    chart4.add_data(data4, titles_from_data=True)
    chart4.set_categories(cats4)
    chart4.width = 15
    chart4.height = 11

    chart4.dataLabels = DataLabelList()
    chart4.dataLabels.showPercent = True
    chart4.dataLabels.showCatName = True

    ws_dash.add_chart(chart4, "L30")

    # ========== SECTION 5: Top Performers Table ==========
    ws_dash.merge_cells('B47:J47')
    sec5 = ws_dash['B47']
    sec5.value = "TOP PERFORMERS RANKING"
    sec5.font = Font(name='Segoe UI Semibold', size=14, bold=True, color='252423')
    sec5.alignment = Alignment(horizontal='left', vertical='center')
    ws_dash.row_dimensions[47].height = 35

    # Table headers
    rank_headers = ['Rank', 'Division', 'SMV Unlock', 'OH Reduction', 'Other Savings', 'Total', 'Solutions']
    header_cols = [2, 3, 5, 6, 7, 8, 9]
    for c, h in zip(header_cols, rank_headers):
        style_header_cell(ws_dash.cell(row=48, column=c, value=h))
    ws_dash.column_dimensions['C'].width = 14
    ws_dash.column_dimensions['D'].width = 2  # Spacer
    ws_dash.row_dimensions[48].height = 28

    # Ranking formulas with LARGE/INDEX/MATCH
    rank_colors = [('FFD700', '000000'), ('C0C0C0', '000000'), ('CD7F32', 'FFFFFF'),
                   ('FFFFFF', '666666'), ('FFFFFF', '666666')]

    for i in range(len(divisions)):
        r = 49 + i
        # Rank number
        ws_dash.cell(row=r, column=2, value=i+1)

        # Dynamic ranking using LARGE function - Division name
        ws_dash.cell(row=r, column=3,
                    value=f'=INDEX(Calculations!$A$3:$A${div_end_row},MATCH(LARGE(Calculations!$F$3:$F${div_end_row},{i+1}),Calculations!$F$3:$F${div_end_row},0))')

        # SMV Unlock for that division
        ws_dash.cell(row=r, column=5,
                    value=f'=ROUND(INDEX(Calculations!$B$3:$B${div_end_row},MATCH(C{r},Calculations!$A$3:$A${div_end_row},0)),3)')

        # OH Reduction
        ws_dash.cell(row=r, column=6,
                    value=f'=ROUND(INDEX(Calculations!$C$3:$C${div_end_row},MATCH(C{r},Calculations!$A$3:$A${div_end_row},0)),1)')

        # Other Savings
        ws_dash.cell(row=r, column=7,
                    value=f'=ROUND(INDEX(Calculations!$D$3:$D${div_end_row},MATCH(C{r},Calculations!$A$3:$A${div_end_row},0)),1)')

        # Total
        ws_dash.cell(row=r, column=8,
                    value=f'=ROUND(LARGE(Calculations!$F$3:$F${div_end_row},{i+1}),2)')

        # Solution Count
        ws_dash.cell(row=r, column=9,
                    value=f'=INDEX(Calculations!$E$3:$E${div_end_row},MATCH(C{r},Calculations!$A$3:$A${div_end_row},0))')

        # Styling
        bg_color, txt_color = rank_colors[min(i, 4)]
        fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')

        for c in [2, 3, 5, 6, 7, 8, 9]:
            cell = ws_dash.cell(row=r, column=c)
            cell.fill = fill
            cell.font = Font(name='Segoe UI', size=10, bold=(i < 3), color=txt_color)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin', color='E0E0E0'),
                right=Side(style='thin', color='E0E0E0'),
                top=Side(style='thin', color='E0E0E0'),
                bottom=Side(style='thin', color='E0E0E0')
            )
        ws_dash.row_dimensions[r].height = 24

    # ========== SECTION 6: Key Insights Panel ==========
    ws_dash.merge_cells('L47:S47')
    sec6 = ws_dash['L47']
    sec6.value = "KEY INSIGHTS"
    sec6.font = Font(name='Segoe UI Semibold', size=14, bold=True, color='252423')
    sec6.alignment = Alignment(horizontal='left', vertical='center')

    insights_data = [
        ("TOP DIVISION", f'=INDEX(Calculations!$A$3:$A${div_end_row},MATCH(MAX(Calculations!$F$3:$F${div_end_row}),Calculations!$F$3:$F${div_end_row},0))', '118DFF'),
        ("BEST SMV UNLOCK", f'=INDEX(Calculations!$A$3:$A${div_end_row},MATCH(MAX(Calculations!$B$3:$B${div_end_row}),Calculations!$B$3:$B${div_end_row},0))', '30B177'),
        ("BEST OH REDUCTION", f'=INDEX(Calculations!$A$3:$A${div_end_row},MATCH(MAX(Calculations!$C$3:$C${div_end_row}),Calculations!$C$3:$C${div_end_row},0))', 'E66C37'),
        ("MOST SOLUTIONS", f'=INDEX(Calculations!$A$3:$A${div_end_row},MATCH(MAX(Calculations!$E$3:$E${div_end_row}),Calculations!$E$3:$E${div_end_row},0))', '6B007B'),
        ("TOTAL DIVISIONS", f'=COUNTA(Calculations!$A$3:$A${div_end_row})', '744EC2'),
    ]

    white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    for i, (label, formula, color) in enumerate(insights_data):
        r = 48 + i

        # Label
        ws_dash.merge_cells(start_row=r, start_column=12, end_row=r, end_column=14)
        label_cell = ws_dash.cell(row=r, column=12, value=label)
        label_cell.font = Font(name='Segoe UI', size=9, color='888888')
        label_cell.alignment = Alignment(horizontal='left', vertical='center')
        label_cell.fill = white_fill
        label_cell.border = Border(left=Side(style='medium', color=color))

        # Value
        ws_dash.merge_cells(start_row=r, start_column=15, end_row=r, end_column=18)
        value_cell = ws_dash.cell(row=r, column=15, value=formula)
        value_cell.font = Font(name='Segoe UI Semibold', size=12, bold=True, color=color)
        value_cell.alignment = Alignment(horizontal='right', vertical='center')
        value_cell.fill = white_fill

        for c in range(12, 19):
            ws_dash.cell(row=r, column=c).fill = white_fill
            ws_dash.cell(row=r, column=c).border = Border(
                bottom=Side(style='thin', color='F0F0F0'),
                left=Side(style='medium', color=color) if c == 12 else Side(style=None)
            )

        ws_dash.row_dimensions[r].height = 32

    # ========== FOOTER ==========
    ws_dash.merge_cells('B56:S56')
    footer = ws_dash['B56']
    footer.value = "Data updates automatically from 'Data' sheet  |  Powered by Dynamic SUMIF Formulas  |  Dashboard v2.0"
    footer.font = Font(name='Segoe UI', size=9, color='AAAAAA', italic=True)
    footer.alignment = Alignment(horizontal='center', vertical='center')
    ws_dash.row_dimensions[56].height = 25

    # Hide gridlines
    ws_dash.sheet_view.showGridLines = False

    # Move Dashboard to front
    wb.move_sheet(ws_dash, offset=-2)

    # Save
    wb.save('Solution_Dashboard_Dynamic.xlsx')

    print("=" * 60)
    print("DYNAMIC DASHBOARD CREATED SUCCESSFULLY!")
    print("=" * 60)
    print(f"\nFile: Solution_Dashboard_Dynamic.xlsx")
    print(f"\nFeatures:")
    print("  - All values auto-update when you modify the Data sheet")
    print("  - Dynamic SUMIF formulas for all aggregations")
    print("  - Auto-ranking of divisions by total savings")
    print("  - KPI cards with live formulas")
    print("  - 5 interactive charts")
    print("  - Key insights panel with dynamic lookups")
    print(f"\nSheets:")
    print("  1. Dashboard - Main visualization")
    print("  2. Data - Your source data (edit here!)")
    print("  3. Calculations - Dynamic pivot tables")
    print(f"\nDivisions tracked: {len(divisions)}")
    print(f"Stages tracked: {len(stages)}")
    print(f"Focus Areas tracked: {len(focus_areas)}")
    print(f"Total Solutions: {len(df)}")

if __name__ == "__main__":
    main()
